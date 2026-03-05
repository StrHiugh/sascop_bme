from django.shortcuts import render
from django.http import JsonResponse, FileResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from core.utils import ejecutar_query_sql, fn_ejecutar_query_sql_lotes, fn_enviar_correo_reporte_bi
from ..models import Frente, Estatus, Paso, PasoOt
import json
import openpyxl
from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import traceback
import io
from django.contrib.auth.decorators import permission_required


@login_required
@permission_required("operaciones.view_centro_consulta", raise_exception=True)
def fn_centro_consulta(request):
   """
   Vista principal para el dashboard de Business Intelligence (BI).
   Renderiza el template estático inicialmente.
   """
   context = {}
   return render(request, "operaciones/centro_consulta/centro_consulta.html", context)

def fn_obtener_subconsulta_origenes(lista_origenes):
   """
   Arma dinámicamente la subconsulta con solo los bloques necesarios.
   """
   bloque_pte = """
      SELECT
         pd.id AS id_origen,
         'PTE' AS tipo,
         COALESCE(ph.oficio_pte, 'SIN FOLIO') AS folio,
         COALESCE(c.descripcion, 'CLIENTE NO ASIGNADO') AS cliente,
         COALESCE(rp.descripcion, 'SIN LÍDER') AS lider,
         'N/A' AS frente,
         NULL::integer AS id_sitio_oficial,
         'NO APLICA' AS sitio_oficial,
         NULL as sitio_pat_desc,
         NULL as sitio_emb_desc,
         NULL as sitio_plat_desc,
         COALESCE(p.descripcion, 'POR DEFINIR') AS documento,
         CASE
            WHEN pd.fecha_entrega IS NULL
               THEN
                     'NO ENTREGADO'
            ELSE TO_CHAR(pd.fecha_entrega, 'DD/MM/YYYY')
         END AS fecha,
         COALESCE(pd.archivo, '') AS archivo,
         pd.fecha_entrega AS _fecha_sort,
         pd.estatus_paso_id AS _fid_estatus_paso,
         ph.id_cliente_id AS _fid_cliente,
         ph.id_responsable_proyecto_id AS _fid_lider,
         NULL::bigint AS _fid_frente,
         NULL::integer AS _fid_patio,
         NULL::integer AS _fid_embarcacion,
         NULL::integer AS _fid_plataforma,
         ce.descripcion AS _descripcion_estatus
      FROM
         pte_detalle pd
      INNER JOIN pte_header ph ON
         pd.id_pte_header_id = ph.id
      LEFT JOIN cliente c ON
         ph.id_cliente_id = c.id
      LEFT JOIN responsable_proyecto rp ON
         ph.id_responsable_proyecto_id = rp.id
      LEFT JOIN paso p ON
         pd.id_paso_id = p.id
      LEFT JOIN cat_estatus ce ON
         pd.estatus_paso_id = ce.id
      WHERE
         ph.estatus != 0
   """

   bloque_ot = """
      SELECT
         od.id AS id_origen,
         'OT' AS tipo,
         COALESCE(o.orden_trabajo, 'SIN OT') AS folio,
         COALESCE(c.descripcion, 'CLIENTE NO ASIGNADO') AS cliente,
         COALESCE(rp.descripcion, 'SIN LÍDER') AS lider,
         COALESCE(f.descripcion, 'SIN FRENTE') AS frente,
         CASE
            WHEN o.id_frente_id = 1
               THEN
                  o.id_patio
            WHEN o.id_frente_id = 2
               THEN
                  o.id_embarcacion
            WHEN o.id_frente_id = 4
               THEN
                  o.id_plataforma
            ELSE
               NULL
         END AS id_sitio_oficial,
         CASE
            WHEN o.id_frente_id = 1
               THEN
                  COALESCE(s_pat.descripcion, 'SIN PATIO')
            WHEN o.id_frente_id = 2
               THEN
                  COALESCE(s_emb.descripcion, 'SIN EMBARCACION')
            WHEN o.id_frente_id = 4
               THEN
                  COALESCE(s_plat.descripcion, 'SIN PLATAFORMA')
            ELSE 'SIN UBICACIÓN'
         END AS sitio_oficial,
         s_pat.descripcion as sitio_pat_desc,
         s_emb.descripcion as sitio_emb_desc,
         s_plat.descripcion as sitio_plat_desc,
         COALESCE(pot.descripcion, 'POR DEFINIR') AS documento,
         CASE
            WHEN od.fecha_entrega IS NULL
               THEN
                  'NO ENTREGADO'
            ELSE
               TO_CHAR(od.fecha_entrega, 'DD/MM/YYYY')
         END AS fecha,
         COALESCE(od.archivo, '') AS archivo,
         od.fecha_entrega AS _fecha_sort,
         od.estatus_paso_id AS _fid_estatus_paso,
         o.id_cliente_id AS _fid_cliente,
         o.id_responsable_proyecto_id AS _fid_lider,
         o.id_frente_id AS _fid_frente,
         o.id_patio AS _fid_patio,
         o.id_embarcacion AS _fid_embarcacion,
         o.id_plataforma AS _fid_plataforma,
         ce.descripcion AS _descripcion_estatus
      FROM
         ot_detalle od
      INNER JOIN ot o ON
         od.id_ot_id = o.id
      LEFT JOIN cliente c ON
         o.id_cliente_id = c.id
      LEFT JOIN responsable_proyecto rp ON
         o.id_responsable_proyecto_id = rp.id
      LEFT JOIN frente f ON
         o.id_frente_id = f.id
      LEFT JOIN paso_ot pot ON
         od.id_paso_id = pot.id
      LEFT JOIN sitio s_pat ON
         o.id_patio = s_pat.id
      LEFT JOIN sitio s_emb ON
         o.id_embarcacion = s_emb.id
      LEFT JOIN sitio s_plat ON
         o.id_plataforma = s_plat.id
      LEFT JOIN cat_estatus ce ON
         od.estatus_paso_id = ce.id
      WHERE
         o.estatus = 1
   """

   bloque_prod_reportes = """
      SELECT
         rmh.id AS id_origen,
         'PROD' AS tipo,
         COALESCE(o.orden_trabajo, 'SIN OT') AS folio,
         COALESCE(c.descripcion, 'CLIENTE NO ASIGNADO') AS cliente,
         COALESCE(rp.descripcion, 'SIN LÍDER') AS lider,
         COALESCE(f.descripcion, 'SIN FRENTE') AS frente,
         CASE
            WHEN o.id_frente_id = 1 THEN o.id_patio
            WHEN o.id_frente_id = 2 THEN o.id_embarcacion
            WHEN o.id_frente_id = 4 THEN o.id_plataforma
            ELSE NULL
         END AS id_sitio_oficial,
         CASE
            WHEN o.id_frente_id = 1 THEN COALESCE(s_pat.descripcion, 'SIN PATIO')
            WHEN o.id_frente_id = 2 THEN COALESCE(s_emb.descripcion, 'SIN EMBARCACION')
            WHEN o.id_frente_id = 4 THEN COALESCE(s_plat.descripcion, 'SIN PLATAFORMA')
            ELSE 'SIN UBICACIÓN'
         END AS sitio_oficial,
         s_pat.descripcion AS sitio_pat_desc,
         s_emb.descripcion AS sitio_emb_desc,
         s_plat.descripcion AS sitio_plat_desc,
         'REPORTE MENSUAL' AS documento,
         CASE
            WHEN rmh.mes IS NOT NULL AND rmh.anio IS NOT NULL
               THEN TO_CHAR(
                  TO_DATE(rmh.anio::text || '-' || LPAD(rmh.mes::text, 2, '0') || '-01', 'YYYY-MM-DD'),
                  'DD/MM/YYYY'
               )
            ELSE 'SIN FECHA'
         END AS fecha,
         COALESCE(rmh.archivo, '') AS archivo,
         CASE
            WHEN rmh.mes IS NOT NULL AND rmh.anio IS NOT NULL
               THEN TO_DATE(rmh.anio::text || '-' || LPAD(rmh.mes::text, 2, '0') || '-01', 'YYYY-MM-DD')
            ELSE NULL
         END AS _fecha_sort,
         rmh.id_estatus_id AS _fid_estatus_paso,
         o.id_cliente_id AS _fid_cliente,
         o.id_responsable_proyecto_id AS _fid_lider,
         o.id_frente_id AS _fid_frente,
         o.id_patio AS _fid_patio,
         o.id_embarcacion AS _fid_embarcacion,
         o.id_plataforma AS _fid_plataforma,
         ce.descripcion AS _descripcion_estatus
      FROM
         reporte_mensual_header rmh
      INNER JOIN ot o ON
         rmh.id_ot_id = o.id
      LEFT JOIN cliente c ON
         o.id_cliente_id = c.id
      LEFT JOIN responsable_proyecto rp ON
         o.id_responsable_proyecto_id = rp.id
      LEFT JOIN frente f ON
         o.id_frente_id = f.id
      LEFT JOIN sitio s_pat ON
         o.id_patio = s_pat.id
      LEFT JOIN sitio s_emb ON
         o.id_embarcacion = s_emb.id
      LEFT JOIN sitio s_plat ON
         o.id_plataforma = s_plat.id
      LEFT JOIN cat_estatus ce ON
         rmh.id_estatus_id = ce.id
      WHERE
         o.estatus = 1
   """

   bloque_prod_gpu = """
      SELECT
         gpu.id AS id_origen,
         'PROD' AS tipo,
         COALESCE(o.orden_trabajo, 'SIN OT') AS folio,
         COALESCE(c.descripcion, 'CLIENTE NO ASIGNADO') AS cliente,
         COALESCE(rp.descripcion, 'SIN LÍDER') AS lider,
         COALESCE(f.descripcion, 'SIN FRENTE') AS frente,
         CASE
            WHEN o.id_frente_id = 1 THEN o.id_patio
            WHEN o.id_frente_id = 2 THEN o.id_embarcacion
            WHEN o.id_frente_id = 4 THEN o.id_plataforma
            ELSE NULL
         END AS id_sitio_oficial,
         CASE
            WHEN o.id_frente_id = 1 THEN COALESCE(s_pat.descripcion, 'SIN PATIO')
            WHEN o.id_frente_id = 2 THEN COALESCE(s_emb.descripcion, 'SIN EMBARCACION')
            WHEN o.id_frente_id = 4 THEN COALESCE(s_plat.descripcion, 'SIN PLATAFORMA')
            ELSE 'SIN UBICACIÓN'
         END AS sitio_oficial,
         s_pat.descripcion AS sitio_pat_desc,
         s_emb.descripcion AS sitio_emb_desc,
         s_plat.descripcion AS sitio_plat_desc,
         'GENERADOR DE PRECIOS UNITARIOS' AS documento,
         CASE
            WHEN p.fecha_produccion IS NULL THEN 'NO ENTREGADO'
            ELSE TO_CHAR(p.fecha_produccion, 'DD/MM/YYYY')
         END AS fecha,
         COALESCE(gpu.archivo, '') AS archivo,
         p.fecha_produccion AS _fecha_sort,
         gpu.id_estatus_id AS _fid_estatus_paso,
         o.id_cliente_id AS _fid_cliente,
         o.id_responsable_proyecto_id AS _fid_lider,
         o.id_frente_id AS _fid_frente,
         o.id_patio AS _fid_patio,
         o.id_embarcacion AS _fid_embarcacion,
         o.id_plataforma AS _fid_plataforma,
         ce.descripcion AS _descripcion_estatus
      FROM
         registro_generadores_pu gpu
      INNER JOIN produccion p ON
         gpu.id_produccion_id = p.id
      INNER JOIN partida_anexo_importada pai ON
         p.id_partida_anexo_id = pai.id
      INNER JOIN importacion_anexo ia ON
         pai.importacion_anexo_id = ia.id
      INNER JOIN ot o ON
         ia.ot_id = o.id
      LEFT JOIN cliente c ON
         o.id_cliente_id = c.id
      LEFT JOIN responsable_proyecto rp ON
         o.id_responsable_proyecto_id = rp.id
      LEFT JOIN frente f ON
         o.id_frente_id = f.id
      LEFT JOIN sitio s_pat ON
         o.id_patio = s_pat.id
      LEFT JOIN sitio s_emb ON
         o.id_embarcacion = s_emb.id
      LEFT JOIN sitio s_plat ON
         o.id_plataforma = s_plat.id
      LEFT JOIN cat_estatus ce ON
         gpu.id_estatus_id = ce.id
      WHERE
         o.estatus = 1
   """

   subconsultas = []
   origenes_validos = lista_origenes if lista_origenes else ["PTE", "OT"]

   if "PTE" in origenes_validos:
      subconsultas.append(bloque_pte)
   if "OT" in origenes_validos:
      subconsultas.append(bloque_ot)
   if "PROD" in origenes_validos:
      subconsultas.append(bloque_prod_reportes)
      subconsultas.append(bloque_prod_gpu)

   return " UNION ALL ".join(subconsultas)


def fn_construir_where_dinamico(filtros):
   """
   Construye dinámicamente la cláusula WHERE y su diccionario de parámetros,
   agregando solo las condiciones que realmente aplican según los filtros recibidos.
   Retorna una tupla (clausula_where_str, params_dict).
   """
   lista_lideres    = filtros.get("lideres_id", [])
   lista_clientes   = filtros.get("clientes_id", [])
   lista_frentes    = filtros.get("frentes_id", [])
   lista_sitios     = filtros.get("sitios_id", [])
   lista_documentos = filtros.get("nombres_doc", [])
   lista_estatus    = filtros.get("estatus_proceso_id", [])
   lista_ots        = filtros.get("ots_id", [])

   fecha_ini_input = filtros.get("fecha_inicio")
   fecha_fin_input = filtros.get("fecha_fin")
   texto_busqueda  = filtros.get("texto_busqueda", "")

   check_entregados = filtros.get("check_entregados")
   check_pendientes = filtros.get("check_no_entregados")

   buscar_por_frente = filtros.get("buscar_por_frente") or "1"

   ninguno_marcado  = not check_entregados and not check_pendientes
   filtro_entregado = check_entregados or ninguno_marcado
   filtro_pendiente = check_pendientes or ninguno_marcado

   condiciones = []
   params = {}

   if lista_ots:
      condiciones.append("T.folio IN %(ids_ots)s")
      params["ids_ots"] = tuple(lista_ots)

   if lista_lideres:
      condiciones.append("T._fid_lider::text IN %(ids_lideres)s")
      params["ids_lideres"] = tuple(lista_lideres)

   if lista_clientes:
      condiciones.append("T._fid_cliente::text IN %(ids_clientes)s")
      params["ids_clientes"] = tuple(lista_clientes)

   if lista_documentos:
      condiciones.append("T.documento IN %(nombres_documentos)s")
      params["nombres_documentos"] = tuple(lista_documentos)

   if lista_estatus:
      condiciones.append("T._fid_estatus_paso::text IN %(ids_estatus)s")
      params["ids_estatus"] = tuple(lista_estatus)

   if texto_busqueda:
      condiciones.append("""
         (
            T.documento ILIKE %(texto)s OR
            T.folio ILIKE %(texto)s OR
            T.cliente ILIKE %(texto)s OR
            T.lider ILIKE %(texto)s
         )
      """)
      params["texto"] = f"%{texto_busqueda}%"

   if lista_sitios:
      params["ids_sitios"] = tuple(lista_sitios)
      if buscar_por_frente == "1":
         if lista_frentes:
            params["ids_frentes"] = tuple(lista_frentes)
            condiciones.append("""
               (
                  T._fid_frente::text IN %(ids_frentes)s AND
                  (
                     (T._fid_frente = 1 AND T._fid_patio::text IN %(ids_sitios)s) OR
                     (T._fid_frente = 2 AND T._fid_embarcacion::text IN %(ids_sitios)s) OR
                     (T._fid_frente = 4 AND T._fid_plataforma::text IN %(ids_sitios)s)
                  )
               )
            """)
         else:
            condiciones.append("""
               (
                  (T._fid_frente = 1 AND T._fid_patio::text IN %(ids_sitios)s) OR
                  (T._fid_frente = 2 AND T._fid_embarcacion::text IN %(ids_sitios)s) OR
                  (T._fid_frente = 4 AND T._fid_plataforma::text IN %(ids_sitios)s)
               )
            """)
      else:
         condiciones.append("""
            (
               T._fid_patio::text IN %(ids_sitios)s OR
               T._fid_embarcacion::text IN %(ids_sitios)s OR
               T._fid_plataforma::text IN %(ids_sitios)s
            )
         """)

   elif lista_frentes:
      condiciones.append("T._fid_frente::text IN %(ids_frentes)s")
      params["ids_frentes"] = tuple(lista_frentes)

   if filtro_entregado and filtro_pendiente:
      pass
   elif filtro_entregado:
      condiciones.append("(T.archivo IS NOT NULL AND LENGTH(TRIM(T.archivo)) > 5)")
   elif filtro_pendiente:
      condiciones.append("(T.archivo IS NULL OR LENGTH(TRIM(T.archivo)) <= 5)")

   if fecha_ini_input and fecha_fin_input:
      condiciones.append("T._fecha_sort BETWEEN %(fecha_ini)s::date AND %(fecha_fin)s::date")
      params["fecha_ini"] = fecha_ini_input
      params["fecha_fin"] = fecha_fin_input

   if condiciones:
      clausula_where = "WHERE\n         " + "\n         AND ".join(condiciones)
   else:
      clausula_where = ""

   params["buscar_por_frente"] = buscar_por_frente
   params["sw_sitio"]          = 1 if lista_sitios else 0
   params["ids_sitios"]        = params.get("ids_sitios", ('-1',))

   return clausula_where, params

@csrf_exempt
@require_http_methods(["POST"])
@login_required
def fn_api_busqueda_global(request):
   """
   Endpoint API para el Buscador Global.
   """
   try:
      cuerpo_peticion = json.loads(request.body)
      salto_registros  = int(cuerpo_peticion.get("start", 0))
      limite_registros = int(cuerpo_peticion.get("length", 10))
      numero_dibujo    = int(cuerpo_peticion.get("draw", 1))

      resultados_paginados, total_filtrados = fn_ejecutar_busqueda_global(request, cuerpo_peticion, salto_registros, limite_registros)

      return JsonResponse({
         "draw": numero_dibujo,
         "recordsTotal": total_filtrados,
         "recordsFiltered": total_filtrados,
         "data": resultados_paginados
      }, status=200)

   except json.JSONDecodeError:
      return JsonResponse({"estatus": "error", "mensaje": "JSON inválido"}, status=400)
   except Exception as error_servidor:
      print(f"Error en Buscador: {str(error_servidor)}")
      return JsonResponse({"estatus": "error", "mensaje": str(error_servidor)}, status=500)


def fn_ejecutar_busqueda_global(request, payload, salto_bd, limite_bd):
   """
   Ejecuta el bloque SQL maestro paginado.
   """
   filtros = payload.get("filtros", {})
   origenes = filtros.get("origenes", [])

   subconsulta_dinamica   = fn_obtener_subconsulta_origenes(origenes)
   clausula_where, params = fn_construir_where_dinamico(filtros)
   params["limite_bd"]    = limite_bd
   params["salto_bd"]     = salto_bd

   sql_conteo = f"""
      SELECT COUNT(*) AS total_registros
         FROM (
      {subconsulta_dinamica}
         ) AS T
      {clausula_where}
   """

   sql_datos = f"""
      SELECT
         T.id_origen,
         T.tipo,
         T.folio,
         T.cliente,
         T.lider,
         T.frente,
         T._fid_estatus_paso AS estatus_paso_id,

         CASE
            WHEN %(buscar_por_frente)s = '0' AND %(sw_sitio)s = 1 AND T._fid_plataforma::text IN %(ids_sitios)s
               THEN T._fid_plataforma
            WHEN %(buscar_por_frente)s = '0' AND %(sw_sitio)s = 1 AND T._fid_embarcacion::text IN %(ids_sitios)s
               THEN T._fid_embarcacion
            WHEN %(buscar_por_frente)s = '0' AND %(sw_sitio)s = 1 AND T._fid_patio::text IN %(ids_sitios)s
               THEN T._fid_patio
            ELSE
               T.id_sitio_oficial
         END AS id_sitio,

         CASE
            WHEN %(buscar_por_frente)s = '1'
               THEN T.sitio_oficial
            WHEN %(buscar_por_frente)s = '0' AND %(sw_sitio)s = 1 AND T._fid_plataforma::text IN %(ids_sitios)s
               THEN T.sitio_plat_desc
            WHEN %(buscar_por_frente)s = '0' AND %(sw_sitio)s = 1 AND T._fid_embarcacion::text IN %(ids_sitios)s
               THEN T.sitio_emb_desc
            WHEN %(buscar_por_frente)s = '0' AND %(sw_sitio)s = 1 AND T._fid_patio::text IN %(ids_sitios)s
               THEN T.sitio_pat_desc
            ELSE CONCAT_WS(' / ',
               NULLIF(T.sitio_pat_desc,  ''),
               NULLIF(T.sitio_emb_desc,  ''),
               NULLIF(T.sitio_plat_desc, '')
            )
         END AS sitio,

         T.documento,
         T.fecha,
         T.archivo
      FROM (
   {subconsulta_dinamica}
      ) AS T
   {clausula_where}
      ORDER BY
         T._fecha_sort ASC NULLS LAST
      LIMIT
         %(limite_bd)s OFFSET %(salto_bd)s;
   """

   resultado_conteo     = ejecutar_query_sql(sql_conteo, params)
   total_filtrados      = resultado_conteo[0]["total_registros"] if resultado_conteo else 0
   resultados_paginados = ejecutar_query_sql(sql_datos, params)

   return resultados_paginados, total_filtrados


@csrf_exempt
@require_http_methods(["POST"])
@login_required
def fn_api_obtener_dashboard(request):
   """
   Endpoint exclusivo para alimentar las gráficas.
   """
   try:
      cuerpo_peticion = json.loads(request.body)
      registros_crudos, modo_sitio_libre = fn_ejecutar_query_graficas(cuerpo_peticion)
      datos_agrupados = fn_agrupar_datos_dashboard(registros_crudos, modo_sitio_libre)

      return JsonResponse({
         "estatus": "ok",
         "mensaje": "Dashboard calculado correctamente",
         "data": datos_agrupados
      }, status=200)

   except json.JSONDecodeError:
      return JsonResponse({"estatus": "error", "mensaje": "Formato JSON inválido"}, status=400)
   except Exception as error_servidor:
      print(f"Error en dashboard: {str(error_servidor)}")
      return JsonResponse({"estatus": "error", "mensaje": "Falla interna del servidor"}, status=500)


def fn_ejecutar_query_graficas(payload):
   """
   Ejecuta la consulta completa para alimentar los dashboards.
   """
   filtros = payload.get("filtros", {})
   origenes = filtros.get("origenes", [])

   subconsulta_dinamica   = fn_obtener_subconsulta_origenes(origenes)
   clausula_where, params = fn_construir_where_dinamico(filtros)

   buscar_por_frente = params["buscar_por_frente"]
   modo_sitio_libre  = (buscar_por_frente == "0" and params["sw_sitio"] == 0)

   sql = f"""
      SELECT
         T.lider,
         T.tipo,
         T.documento,
         T.folio,
         T.frente,
         T.cliente,
         T._fid_estatus_paso AS estatus_paso_id,
         T.sitio_pat_desc,
         T.sitio_emb_desc,
         T.sitio_plat_desc,
         T.sitio_oficial,
         CASE
            WHEN %(buscar_por_frente)s = '1'
               THEN T.sitio_oficial
            WHEN %(buscar_por_frente)s = '0' AND %(sw_sitio)s = 1 AND T._fid_plataforma::text IN %(ids_sitios)s
               THEN T.sitio_plat_desc
            WHEN %(buscar_por_frente)s = '0' AND %(sw_sitio)s = 1 AND T._fid_embarcacion::text IN %(ids_sitios)s
               THEN T.sitio_emb_desc
            WHEN %(buscar_por_frente)s = '0' AND %(sw_sitio)s = 1 AND T._fid_patio::text IN %(ids_sitios)s
               THEN T.sitio_pat_desc
            ELSE
               T.sitio_oficial
         END AS sitio,

         CASE
            WHEN T.archivo IS NOT NULL AND LENGTH(TRIM(T.archivo)) > 5
               THEN 1
               ELSE 0
         END AS tiene_archivo,
         T._descripcion_estatus
         FROM (
      {subconsulta_dinamica}
         ) AS T
      {clausula_where}
   """

   return fn_ejecutar_query_sql_lotes(sql, params), modo_sitio_libre


@login_required
def fn_obtener_frente_afectacion_dos(solicitud):
   try:
      frentes_listado = Frente.objects.filter(nivel_afectacion=2, activo=True)
      return JsonResponse([
         {
            "id": f.id,
            "descripcion": f.descripcion,
            "afectacion": f.nivel_afectacion,
            "comentario": f.comentario,
            "activo": f.activo
         } for f in frentes_listado
      ], safe=False) if frentes_listado.exists() else JsonResponse({
         "tipo_aviso": "error",
         "detalles": "No se encontraron frentes con afectación 2"
      }, status=404)
   except Exception as error_proceso:
      return JsonResponse({"tipo_aviso": "error", "detalles": f"Error en el servidor: {str(error_proceso)}"}, status=500)


@login_required
def fn_obtener_estatus_afectacion_uno(request):
   try:
      estatus_listado = Estatus.objects.filter(nivel_afectacion=1, activo=True)
      if not estatus_listado.exists():
         return JsonResponse({
            "tipo_aviso": "error",
            "detalles": "No se encontraron estatus con nivel de afectacion uno"
         }, status=404)
      data = [
         {
            "id": f.id,
            "descripcion": f.descripcion,
            "nivel_afectacion": f.nivel_afectacion,
            "comentario": f.comentario,
            "activo": f.activo
         } for f in estatus_listado
      ]
      return JsonResponse(data, safe=False)
   except Exception as error_proceso:
      return JsonResponse({"tipo_aviso": "error", "detalles": f"Error en el servidor: {str(error_proceso)}"}, status=500)


@login_required
def fn_obtener_catalogo_documentos_unificado(request):
   try:
      qs_pte = Paso.objects.filter(activo=True).values("descripcion")
      qs_ot  = PasoOt.objects.filter(activo=True).values("descripcion")
      consulta_unificada = qs_pte.union(qs_ot).order_by("descripcion")
      resultados = [
         {
            "id": fila["descripcion"],
            "descripcion": fila["descripcion"]
         }
         for fila in consulta_unificada
      ]
      return JsonResponse(resultados, safe=False)
   except Exception as e:
      print(f"Error obteniendo catálogo unificado: {str(e)}")
      return JsonResponse([], safe=False)

def fn_agrupar_datos_dashboard(registros_db, modo_sitio_libre=False):
   totales = {"cargados": 0, "pendientes": 0, "no_aplica": 0}
   origenes = {"PTE": 0, "OT": 0, "PROD": 0}
   lideres = {}
   documentos = {}
   clientes = {}
   estatus_embudo = {}
   frentes = {}
   sitios = {}
   folios = {}

   for fila in registros_db:
      lider = fila.get("lider", "SIN LÍDER")
      tipo = fila.get("tipo", "DESCONOCIDO")
      documento = fila.get("documento", "SIN DOCUMENTO")
      folio = fila.get("folio", "SIN FOLIO")
      frente = fila.get("frente", "SIN FRENTE")
      sitio = fila.get("sitio", "SIN SITIO")
      cliente = fila.get("cliente", "SIN CLIENTE")
      descripcion_estatus = fila.get("_descripcion_estatus")
      texto_estatus = descripcion_estatus if descripcion_estatus else "ESTATUS DESCONOCIDO"
      estatus_id = fila.get("estatus_paso_id")

      valor_archivo = fila.get("tiene_archivo", 0)
      es_cargado = True if valor_archivo == 1 else False
      es_no_aplica = True if (estatus_id == 14 or documento == "NO APLICA") else False

      if tipo in origenes:
         origenes[tipo] += 1

      llave_estatus = texto_estatus
      estatus_embudo[llave_estatus] = estatus_embudo.get(llave_estatus, 0) + 1

      if tipo in ["OT", "PROD"]:
         frentes[frente] = frentes.get(frente, 0) + 1
         if modo_sitio_libre:
            for desc_sitio in [
                  fila.get("sitio_pat_desc"),
                  fila.get("sitio_emb_desc"),
                  fila.get("sitio_plat_desc"),
            ]:
                  if desc_sitio:
                     sitios[desc_sitio] = sitios.get(desc_sitio, 0) + 1
         else:
            sitio_resuelto = fila.get("sitio", "SIN SITIO")
            sitios[sitio_resuelto] = sitios.get(sitio_resuelto, 0) + 1

      if es_no_aplica:
         totales["no_aplica"] += 1
      elif es_cargado:
         totales["cargados"] += 1
      else:
         totales["pendientes"] += 1

      for tabla, clave in [(lideres, lider), (documentos, documento), (clientes, cliente), (folios, folio)]:
         if clave not in tabla:
            tabla[clave] = {"cargados": 0, "pendientes": 0, "no_aplica": 0}
         
         if es_no_aplica:
            tabla[clave]["no_aplica"] += 1
         elif es_cargado:
            tabla[clave]["cargados"] += 1
         else:
            tabla[clave]["pendientes"] += 1

   datos_procesados = {
      "totales_generales": totales,
      "distribucion_origenes": [{"origen": llave, "total": valor} for llave, valor in origenes.items()],
      "rendimiento_lideres": [{"nombre": llave, "cargados": valor["cargados"], "pendientes": valor["pendientes"], "no_aplica": valor["no_aplica"]} for llave, valor in lideres.items()],
      "tipos_documentos": [{"documento": llave, "cargados": valor["cargados"], "pendientes": valor["pendientes"], "no_aplica": valor["no_aplica"]} for llave, valor in documentos.items()],
      "estatus_clientes": [{"cliente": llave, "cargados": valor["cargados"], "pendientes": valor["pendientes"], "no_aplica": valor["no_aplica"]} for llave, valor in clientes.items()],
      "embudo_estatus": [{"estatus": llave, "total": valor} for llave, valor in estatus_embudo.items()],
      "frentes_ot": [{"frente": llave, "total": valor} for llave, valor in frentes.items()],
      "sitios_ot": [{"sitio": llave, "total": valor} for llave, valor in sitios.items()],
      "avance_folios": [{"folio": llave, "cargados": valor["cargados"], "pendientes": valor["pendientes"], "no_aplica": valor["no_aplica"]} for llave, valor in folios.items()]
   }

   return datos_procesados

@csrf_exempt
@require_http_methods(["POST"])
@login_required
def fn_api_descargar_excel_bi(request):
   try:
      cuerpo_peticion = json.loads(request.body)
      filtros = cuerpo_peticion.get("filtros", {})
      origenes = filtros.get("origenes", [])

      subconsulta_dinamica = fn_obtener_subconsulta_origenes(origenes)
      clausula_where, params = fn_construir_where_dinamico(filtros)

      sql_excel = f"""
         SELECT
            T.tipo AS "Origen",
            T.folio AS "Folio",
            T.cliente AS "Cliente",
            T.lider AS "Líder",
            T.frente AS "Frente",
            CASE
               WHEN %(buscar_por_frente)s = '1' THEN T.sitio_oficial
               WHEN %(buscar_por_frente)s = '0' AND %(sw_sitio)s = 1 AND T._fid_plataforma::text IN %(ids_sitios)s THEN T.sitio_plat_desc
               WHEN %(buscar_por_frente)s = '0' AND %(sw_sitio)s = 1 AND T._fid_embarcacion::text IN %(ids_sitios)s THEN T.sitio_emb_desc
               WHEN %(buscar_por_frente)s = '0' AND %(sw_sitio)s = 1 AND T._fid_patio::text IN %(ids_sitios)s THEN T.sitio_pat_desc
               ELSE CONCAT_WS(' / ', NULLIF(T.sitio_pat_desc, ''), NULLIF(T.sitio_emb_desc, ''), NULLIF(T.sitio_plat_desc, ''))
            END AS "Sitio",
            T.documento AS "Documento",
            T._descripcion_estatus AS "Estatus",
            CASE
               WHEN T.archivo IS NOT NULL AND LENGTH(TRIM(T.archivo)) > 5 THEN 'ENTREGADO'
               WHEN T._fid_estatus_paso = 14 THEN 'NO APLICA'
               ELSE 'PENDIENTE'
            END AS "Estado Archivo",
            T.archivo AS "Enlace Documento",
            T.fecha AS "Fecha Modificación"
         FROM (
      {subconsulta_dinamica}
         ) AS T
      {clausula_where}
         ORDER BY
            T._fecha_sort ASC NULLS LAST;
      """

      registros = fn_ejecutar_query_sql_lotes(sql_excel, params)

      libro_trabajo = openpyxl.Workbook(write_only=True)
      hoja_pte = libro_trabajo.create_sheet(title="Reporte PTE")
      hoja_ot = libro_trabajo.create_sheet(title="Reporte OT")
      hoja_prod = libro_trabajo.create_sheet(title="Reporte PROD")

      diccionario_hojas = {"PTE": hoja_pte, "OT": hoja_ot, "PROD": hoja_prod}
      conteos_hojas = {"PTE": 0, "OT": 0, "PROD": 0}

      estilo_cabecera = PatternFill(start_color="F05523", end_color="F05523", fill_type="solid")
      fuente_cabecera = Font(color="FFFFFF", bold=True)
      fuente_enlace = Font(color="0563C1", underline="single")
      alineacion_centro = Alignment(horizontal="center", vertical="center")
      alineacion_izquierda = Alignment(horizontal="left", vertical="center")
      borde_delgado = Border(
         left=Side(style="thin", color="D0D1D3"),
         right=Side(style="thin", color="D0D1D3"),
         top=Side(style="thin", color="D0D1D3"),
         bottom=Side(style="thin", color="D0D1D3")
      )

      columnas = [
         "Origen", "Folio", "Cliente", "Líder", "Frente", "Sitio",
         "Documento", "Estatus", "Estado Archivo", "Enlace Documento", "Fecha Modificación"
      ]

      anchos_minimos = {
         "Folio": 18, "Cliente": 35, "Líder": 30, "Frente": 20,
         "Sitio": 25, "Documento": 40, "Estatus": 15, "Enlace Documento": 18
      }

      for nombre_hoja, objeto_hoja in diccionario_hojas.items():
         for indice, col_nombre in enumerate(columnas, 1):
            letra_columna = get_column_letter(indice)
            objeto_hoja.column_dimensions[letra_columna].width = anchos_minimos.get(col_nombre, 15)

         fila_cabecera = []
         for nombre_columna in columnas:
            celda = WriteOnlyCell(objeto_hoja, value=nombre_columna)
            celda.fill = estilo_cabecera
            celda.font = fuente_cabecera
            celda.alignment = alineacion_centro
            celda.border = borde_delgado
            fila_cabecera.append(celda)

         objeto_hoja.append(fila_cabecera)

      dominio_base = request.build_absolute_uri("/")[:-1]

      for fila_datos in registros:
         origen_registro = fila_datos.get("Origen", "")
         hoja_destino = diccionario_hojas.get(origen_registro)

         if not hoja_destino:
            continue

         fila_excel = []

         for nombre_columna in columnas:
            valor_bd = fila_datos.get(nombre_columna, "")
            celda = WriteOnlyCell(hoja_destino, value=valor_bd)

            celda.border = borde_delgado
            celda.alignment = alineacion_izquierda

            if nombre_columna in ["Origen", "Estado Archivo", "Fecha Modificación"]:
               celda.alignment = alineacion_centro
               celda.value = valor_bd

            elif nombre_columna == "Enlace Documento":
               if valor_bd and len(str(valor_bd).strip()) > 5:
                  url_completa = valor_bd if valor_bd.startswith("http") else f"{dominio_base}{valor_bd}"
                  celda.value = "Abrir Archivo"
                  celda.hyperlink = url_completa
                  celda.font = fuente_enlace
                  celda.alignment = alineacion_centro
               else:
                  celda.value = "Sin Archivo"
                  celda.alignment = alineacion_centro
            else:
               celda.value = valor_bd

            fila_excel.append(celda)

         hoja_destino.append(fila_excel)
         conteos_hojas[origen_registro] += 1

      if conteos_hojas["PTE"] == 0:
         libro_trabajo.remove(hoja_pte)
      if conteos_hojas["OT"] == 0:
         libro_trabajo.remove(hoja_ot)
      if conteos_hojas["PROD"] == 0:
         libro_trabajo.remove(hoja_prod)

      if not libro_trabajo.sheetnames:
         hoja_vacia = libro_trabajo.create_sheet(title="Sin Datos")
         celda_vacia = WriteOnlyCell(hoja_vacia, value="No se encontraron registros para tu búsqueda.")
         hoja_vacia.append([celda_vacia])

      flujo_memoria = io.BytesIO()
      libro_trabajo.save(flujo_memoria)
      flujo_memoria.seek(0)

      fecha_actual = datetime.now().strftime("%Y%m%d_%H%M")
      nombre_archivo = f"Reporte_SASCOP_BI_{fecha_actual}.xlsx"

      respuesta = FileResponse(
         flujo_memoria,
         as_attachment=True,
         filename=nombre_archivo,
         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
      )

      return respuesta

   except Exception as error_servidor:
      print("Error crítico al generar Excel masivo:")
      traceback.print_exc()
      return JsonResponse({"estatus": "error", "mensaje": "Falla al generar el archivo Excel."}, status=500)

@csrf_exempt
@require_http_methods(["POST"])
@login_required
def fn_api_enviar_correo_bi(request):
   """
   Endpoint para procesar filtros, generar el Excel en memoria validando pesos,
   y enviar el reporte visual por correo electrónico.
   """
   try:
      cuerpo_peticion = json.loads(request.body)
      cadena_correos = cuerpo_peticion.get("correos", "")
      filtros_front = cuerpo_peticion.get("filtros", {})
      graficas_front = cuerpo_peticion.get("graficas", [])

      lista_destinatarios = [correo.strip() for correo in cadena_correos.split(",") if correo.strip()]
      es_valida_lista = True if len(lista_destinatarios) > 0 else False

      if not es_valida_lista:
         return JsonResponse({"estatus": "error", "mensaje": "Correos no válidos."}, status=400)

      origenes = filtros_front.get("origenes", [])
      subconsulta_dinamica = fn_obtener_subconsulta_origenes(origenes)
      clausula_where, params = fn_construir_where_dinamico(filtros_front)

      sql_conteo = f"SELECT COUNT(*) AS total FROM ({subconsulta_dinamica}) AS T {clausula_where}"
      resultado_conteo = ejecutar_query_sql(sql_conteo, params)
      total_registros = resultado_conteo[0]["total"] if resultado_conteo else 0

      tupla_excel = None
      mensaje_advertencia = ""
      supera_registros = True if total_registros > 100000 else False

      if supera_registros:
         mensaje_advertencia = "El volumen de datos supera los 100,000 registros permitidos para envío por correo. Por favor, descarga el Excel masivo directamente desde el sistema."
      else:
         sql_excel = f"""
            SELECT
               T.tipo AS "Origen", T.folio AS "Folio", T.cliente AS "Cliente",
               T.lider AS "Líder", T.frente AS "Frente",
               CASE
                  WHEN %(buscar_por_frente)s = '1' THEN T.sitio_oficial
                  WHEN %(buscar_por_frente)s = '0' AND %(sw_sitio)s = 1 AND T._fid_plataforma::text IN %(ids_sitios)s THEN T.sitio_plat_desc
                  WHEN %(buscar_por_frente)s = '0' AND %(sw_sitio)s = 1 AND T._fid_embarcacion::text IN %(ids_sitios)s THEN T.sitio_emb_desc
                  WHEN %(buscar_por_frente)s = '0' AND %(sw_sitio)s = 1 AND T._fid_patio::text IN %(ids_sitios)s THEN T.sitio_pat_desc
                  ELSE CONCAT_WS(' / ', NULLIF(T.sitio_pat_desc, ''), NULLIF(T.sitio_emb_desc, ''), NULLIF(T.sitio_plat_desc, ''))
               END AS "Sitio",
               T.documento AS "Documento", T._descripcion_estatus AS "Estatus",
               CASE
                  WHEN T.archivo IS NOT NULL AND LENGTH(TRIM(T.archivo)) > 5 THEN 'ENTREGADO'
                  WHEN T._fid_estatus_paso = 14 THEN 'NO APLICA'
                  ELSE 'PENDIENTE'
               END AS "Estado Archivo",
               T.archivo AS "Enlace Documento", T.fecha AS "Fecha Modificación"
            FROM ({subconsulta_dinamica}) AS T {clausula_where}
            ORDER BY T._fecha_sort ASC NULLS LAST;
         """

         registros_excel = fn_ejecutar_query_sql_lotes(sql_excel, params)
         libro_trabajo = openpyxl.Workbook(write_only=True)
         hoja_unica = libro_trabajo.create_sheet(title="Reporte Datos")

         estilo_cabecera = PatternFill(start_color="F05523", end_color="F05523", fill_type="solid")
         fuente_cabecera = Font(color="FFFFFF", bold=True)
         alineacion_centro = Alignment(horizontal="center", vertical="center")

         columnas = ["Origen", "Folio", "Cliente", "Líder", "Frente", "Sitio", "Documento", "Estatus", "Estado Archivo", "Enlace Documento", "Fecha Modificación"]

         fila_cabecera = []
         for col_nombre in columnas:
            celda = WriteOnlyCell(hoja_unica, value=col_nombre)
            celda.fill = estilo_cabecera
            celda.font = fuente_cabecera
            celda.alignment = alineacion_centro
            fila_cabecera.append(celda)

         hoja_unica.append(fila_cabecera)

         for fila_datos in registros_excel:
            fila_excel = []
            for col in columnas:
               celda = WriteOnlyCell(hoja_unica, value=fila_datos.get(col, ""))
               fila_excel.append(celda)
            hoja_unica.append(fila_excel)

         flujo_memoria = io.BytesIO()
         libro_trabajo.save(flujo_memoria)

         peso_bytes = flujo_memoria.getbuffer().nbytes
         peso_mb = peso_bytes / (1024 * 1024)
         supera_peso = True if peso_mb > 20 else False

         if supera_peso:
            mensaje_advertencia = f"El archivo Excel generado pesa {peso_mb:.2f} MB, superando el límite de seguridad de 20 MB. Por favor, descárgalo desde el sistema."
         else:
            fecha_actual = datetime.now().strftime("%Y%m%d_%H%M")
            nombre_archivo = f"Reporte_SASCOP_BI_{fecha_actual}.xlsx"
            tupla_excel = (nombre_archivo, flujo_memoria.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

      envio_exitoso = fn_enviar_correo_reporte_bi(lista_destinatarios, graficas_front, tupla_excel, mensaje_advertencia)

      estatus_respuesta = "ok" if envio_exitoso else "error"
      mensaje_respuesta = "Reporte enviado exitosamente." if envio_exitoso else "Hubo un error al intentar enviar el correo."

      return JsonResponse({"estatus": estatus_respuesta, "mensaje": mensaje_respuesta})

   except Exception as error_endpoint:
      traceback.print_exc()
      return JsonResponse({"estatus": "error", "mensaje": "Ocurrió un error interno procesando la solicitud."}, status=500)

@login_required
def fn_api_obtener_ots_cc(request):
   try:
      q      = request.GET.get("q", "").strip()
      pagina = int(request.GET.get("page", 1))

      if len(q) < 2:
         return JsonResponse({"results": [], "more": False})

      por_pagina = 20
      salto      = (pagina - 1) * por_pagina
      termino    = f"%{q}%"

      sql = """
         SELECT DISTINCT
            o.orden_trabajo AS id,
            o.orden_trabajo AS texto
         FROM ot o
         WHERE o.estatus = 1
            AND o.orden_trabajo ILIKE %(termino)s
         ORDER BY o.orden_trabajo
         LIMIT %(limite)s OFFSET %(salto)s
      """
      params     = {"termino": termino, "limite": por_pagina + 1, "salto": salto}
      resultados = ejecutar_query_sql(sql, params)

      hay_mas   = len(resultados) > por_pagina
      listado   = resultados[:por_pagina]
      respuesta = [{"id": r["id"], "text": r["texto"]} for r in listado]

      return JsonResponse({"results": respuesta, "more": hay_mas})
   except Exception as error_proceso:
      print(f"Error al obtener catálogo de OTs: {str(error_proceso)}")
      return JsonResponse({"results": [], "more": False})

@login_required
def fn_api_obtener_estatus_cobro_cc(request):
   try:
      sql = """
         SELECT DISTINCT
            ce.id,
            ce.descripcion
         FROM cat_estatus ce
         INNER JOIN produccion p ON p.id_estatus_cobro_id = ce.id
         ORDER BY ce.descripcion
      """
      resultados = ejecutar_query_sql(sql, {})
      return JsonResponse(resultados, safe=False)
   except Exception as error_proceso:
      print(f"Error al obtener catálogo de estatus cobro: {str(error_proceso)}")
      return JsonResponse([], safe=False)

@login_required
def fn_api_obtener_anexos_cc(request):
   try:
      sql = """
         SELECT DISTINCT
            pai.anexo AS id,
            pai.anexo AS descripcion
         FROM partida_anexo_importada pai
         WHERE pai.anexo IS NOT NULL AND pai.anexo != ''
         ORDER BY pai.anexo
      """
      resultados = ejecutar_query_sql(sql, {})
      return JsonResponse(resultados, safe=False)
   except Exception as error_proceso:
      print(f"Error al obtener catálogo de anexos: {str(error_proceso)}")
      return JsonResponse([], safe=False)

@login_required
def fn_api_buscar_partidas_cc(request):
   try:
      q      = request.GET.get("q", "").strip()
      pagina = int(request.GET.get("page", 1))

      if len(q) < 2:
         return JsonResponse({"results": [], "more": False})

      por_pagina = 20
      salto      = (pagina - 1) * por_pagina
      termino    = f"%{q}%"

      sql = """
         SELECT DISTINCT
            pai.id_partida AS id,
            pai.id_partida || ' — ' || COALESCE(pai.descripcion_concepto, '') AS texto
         FROM partida_anexo_importada pai
         WHERE
            pai.id_partida IS NOT NULL
            AND (
               pai.id_partida ILIKE %(termino)s
               OR pai.descripcion_concepto ILIKE %(termino)s
            )
         ORDER BY pai.id_partida
         LIMIT %(limite)s OFFSET %(salto)s
      """
      params     = {"termino": termino, "limite": por_pagina + 1, "salto": salto}
      resultados = ejecutar_query_sql(sql, params)

      hay_mas   = len(resultados) > por_pagina
      listado   = resultados[:por_pagina]
      respuesta = [{"id": r["id"], "text": r["texto"]} for r in listado]

      return JsonResponse({"results": respuesta, "more": hay_mas})
   except Exception as error_proceso:
      print(f"Error al buscar partidas: {str(error_proceso)}")
      return JsonResponse({"results": [], "more": False})

@csrf_exempt
@require_http_methods(["POST"])
@login_required
def fn_api_busqueda_prod_informacion(request):
   try:
      cuerpo_peticion  = json.loads(request.body)
      salto_registros  = int(cuerpo_peticion.get("start", 0))
      limite_registros = int(cuerpo_peticion.get("length", 10))
      numero_dibujo    = int(cuerpo_peticion.get("draw", 1))
      filtros          = cuerpo_peticion.get("filtros", {})

      resultados_paginados, total_filtrados, datos_dashboard = fn_ejecutar_query_prod_info(filtros, salto_registros, limite_registros)

      return JsonResponse({
         "draw":            numero_dibujo,
         "recordsTotal":    total_filtrados,
         "recordsFiltered": total_filtrados,
         "data":            resultados_paginados,
         "dashboard":       datos_dashboard,
      }, status=200)

   except json.JSONDecodeError:
      return JsonResponse({"estatus": "error", "mensaje": "JSON inválido"}, status=400)
   except Exception as error_servidor:
      print(f"Error en búsqueda producción info: {str(error_servidor)}")
      return JsonResponse({"estatus": "error", "mensaje": "Falla interna del servidor"}, status=500)

def fn_ejecutar_query_prod_info(filtros, salto_bd, limite_bd):
   lista_ots          = filtros.get("ots_id", [])
   lista_tipos_tiempo = filtros.get("tipos_tiempo", [])
   lista_anexos       = filtros.get("anexos", [])
   lista_clientes     = filtros.get("clientes_id", [])
   lista_lideres      = filtros.get("lideres_id", [])
   lista_partidas     = filtros.get("partidas_id", [])
   lista_sitios       = filtros.get("sitios_id", [])
   es_excedente       = filtros.get("es_excedente")
   texto_busqueda     = filtros.get("texto_busqueda", "")
   fecha_ini_input    = filtros.get("fecha_inicio")
   fecha_fin_input    = filtros.get("fecha_fin")

   condiciones_a = ["o.estatus = 1"]
   condiciones_b = ["o.estatus = 1"]
   params = {}

   if lista_ots:
      condiciones_a.append("o.orden_trabajo IN %(ids_ots)s")
      condiciones_b.append("o.orden_trabajo IN %(ids_ots)s")
      params["ids_ots"] = tuple(lista_ots)

   if lista_clientes:
      condiciones_a.append("o.id_cliente_id::text IN %(ids_clientes)s")
      condiciones_b.append("o.id_cliente_id::text IN %(ids_clientes)s")
      params["ids_clientes"] = tuple(lista_clientes)

   if lista_lideres:
      condiciones_a.append("o.id_responsable_proyecto_id::text IN %(ids_lideres)s")
      condiciones_b.append("o.id_responsable_proyecto_id::text IN %(ids_lideres)s")
      params["ids_lideres"] = tuple(lista_lideres)

   if lista_partidas:
      condiciones_a.append("pai.id_partida IN %(ids_partidas)s")
      condiciones_b.append("pai.id_partida IN %(ids_partidas)s")
      params["ids_partidas"] = tuple(lista_partidas)

   if lista_tipos_tiempo:
      condiciones_a.append("p.tipo_tiempo IN %(tipos_tiempo)s")
      params["tipos_tiempo"] = tuple(lista_tipos_tiempo)

   if lista_anexos:
      condiciones_a.append("pai.anexo IN %(ids_anexos)s")
      condiciones_b.append("pai.anexo IN %(ids_anexos)s")
      params["ids_anexos"] = tuple(lista_anexos)

   if lista_sitios:
      condiciones_a.append("p.id_sitio_produccion_id::text IN %(ids_sitios_info)s")
      condiciones_b.append("FALSE")
      params["ids_sitios_info"] = tuple(lista_sitios)

   if es_excedente is True:
      condiciones_a.append("p.es_excedente = true")
      condiciones_b.append("FALSE")

   if fecha_ini_input and fecha_fin_input:
      condiciones_a.append("p.fecha_produccion BETWEEN %(fecha_ini)s::date AND %(fecha_fin)s::date")
      condiciones_b.append("pp.fecha BETWEEN %(fecha_ini)s::date AND %(fecha_fin)s::date")
      params["fecha_ini"] = fecha_ini_input
      params["fecha_fin"] = fecha_fin_input

   if texto_busqueda:
      cond_texto = """
         (
            pai.id_partida ILIKE %(texto)s OR
            pai.descripcion_concepto ILIKE %(texto)s OR
            o.orden_trabajo ILIKE %(texto)s
         )
      """
      condiciones_a.append(cond_texto)
      condiciones_b.append(cond_texto)
      params["texto"] = f"%{texto_busqueda}%"

   where_a = "WHERE " + " AND ".join(condiciones_a)
   where_b = "WHERE " + " AND ".join(condiciones_b) + """
      AND NOT EXISTS (
         SELECT 1 FROM produccion p2
         WHERE p2.id_partida_anexo_id = pp.partida_anexo_id
         AND p2.fecha_produccion = pp.fecha
      )
   """

   params["limite_bd"] = limite_bd
   params["salto_bd"]  = salto_bd

   sql_union = f"""
      SELECT
         o.orden_trabajo AS ot,
         COALESCE(pai.anexo, '—') AS anexo,
         pai.id_partida || ' — ' || COALESCE(pai.descripcion_concepto, '') AS partida,
         ROUND(p.volumen_produccion::numeric, 4) AS vol_producido,
         ROUND(pai.volumen_proyectado::numeric, 4) AS vol_proyectado,
         ROUND(pp.volumen_programado::numeric, 4) AS vol_programado,
         ROUND((p.volumen_produccion * COALESCE(pai.precio_unitario_mn, 0))::numeric, 2) AS importe_producido,
         ROUND((COALESCE(pp.volumen_programado, 0) * COALESCE(pai.precio_unitario_mn, 0))::numeric, 2) AS importe_programado,
         COALESCE(rp.descripcion, 'SIN LÍDER') AS lider_proyecto,
         COALESCE(um.clave, '—') AS unidad_medida,
         TO_CHAR(p.fecha_produccion, 'DD/MM/YYYY') AS fecha_produccion,
         p.tipo_tiempo AS tipo_tiempo,
         COALESCE(s.descripcion, 'SIN SITIO') AS sitio,
         p.fecha_produccion AS _fecha_ord
      FROM produccion p
      INNER JOIN partida_anexo_importada pai ON p.id_partida_anexo_id = pai.id
      INNER JOIN importacion_anexo ia ON pai.importacion_anexo_id = ia.id and ia.es_activo = true
      INNER JOIN ot o ON ia.ot_id = o.id AND o.estatus = 1
      LEFT JOIN sitio s ON p.id_sitio_produccion_id = s.id AND s.activo = true
      LEFT JOIN partida_proyectada pp
         ON pp.partida_anexo_id = pai.id
         AND pp.fecha = p.fecha_produccion
         AND pp.ot_id = o.id
      LEFT JOIN responsable_proyecto rp ON o.id_responsable_proyecto_id = rp.id
      LEFT JOIN unidad_medida um ON pai.unidad_medida_id = um.id
      {where_a}

      UNION ALL

      SELECT
         o.orden_trabajo AS ot,
         COALESCE(pai.anexo, '—') AS anexo,
         pai.id_partida || ' — ' || COALESCE(pai.descripcion_concepto, '') AS partida,
         0.0000 AS vol_producido,
         ROUND(pai.volumen_proyectado::numeric, 4) AS vol_proyectado,
         ROUND(pp.volumen_programado::numeric, 4) AS vol_programado,
         0.00 AS importe_producido,
         ROUND((pp.volumen_programado * COALESCE(pai.precio_unitario_mn, 0))::numeric, 2) AS importe_programado,
         COALESCE(rp.descripcion, 'SIN LÍDER') AS lider_proyecto,
         COALESCE(um.clave, '—') AS unidad_medida,
         TO_CHAR(pp.fecha, 'DD/MM/YYYY') AS fecha_produccion,
         NULL::varchar(3) AS tipo_tiempo,
         'SIN SITIO' AS sitio,
         pp.fecha AS _fecha_ord
      FROM partida_proyectada pp
      INNER JOIN partida_anexo_importada pai ON pp.partida_anexo_id = pai.id
      INNER JOIN importacion_anexo ia ON pai.importacion_anexo_id = ia.id AND ia.es_activo = true
      INNER JOIN ot o ON pp.ot_id = o.id
      LEFT JOIN responsable_proyecto rp ON o.id_responsable_proyecto_id = rp.id
      LEFT JOIN unidad_medida um ON pai.unidad_medida_id = um.id
      {where_b}
   """

   where_outer = ""

   sql_conteo = f"""
      SELECT COUNT(*) AS total_registros
      FROM ({sql_union}) AS combinado
      {where_outer}
   """

   sql_datos = f"""
      SELECT ot, anexo, partida, vol_producido, vol_proyectado, vol_programado,
             importe_producido, importe_programado, lider_proyecto, unidad_medida,
             fecha_produccion, tipo_tiempo, sitio
      FROM ({sql_union}) AS combinado
      {where_outer}
      ORDER BY _fecha_ord DESC NULLS LAST, partida ASC
      LIMIT %(limite_bd)s OFFSET %(salto_bd)s
   """

   resultado_conteo     = ejecutar_query_sql(sql_conteo, params)
   total_filtrados      = resultado_conteo[0]["total_registros"] if resultado_conteo else 0
   resultados_paginados = ejecutar_query_sql(sql_datos, params)
   datos_dashboard      = fn_calcular_aggregados_info(sql_union, params)

   return resultados_paginados, total_filtrados, datos_dashboard


def fn_calcular_aggregados_info(sql_union, params):
   try:
      sql_resumen = f"""
         SELECT
            COALESCE(SUM(importe_producido), 0)                                   AS total_importe_producido,
            COUNT(DISTINCT CASE WHEN vol_producido > 0 THEN ot END)               AS proyectos_ejecutados,
            COUNT(DISTINCT CASE WHEN vol_producido > 0 THEN fecha_produccion END) AS dias_unicos
         FROM ({sql_union}) AS c
      """

      sql_mejor_dia = f"""
         SELECT fecha_produccion, SUM(importe_producido) AS importe_dia
         FROM ({sql_union}) AS c
         WHERE vol_producido > 0
         GROUP BY fecha_produccion
         ORDER BY importe_dia DESC
         LIMIT 1
      """

      sql_por_fecha = f"""
         SELECT
            fecha_produccion,
            SUM(importe_producido)  AS importe_producido,
            SUM(importe_programado) AS importe_programado
         FROM ({sql_union}) AS c
         GROUP BY fecha_produccion, _fecha_ord
         ORDER BY _fecha_ord ASC NULLS LAST
      """

      sql_por_tipo = f"""
         SELECT tipo_tiempo, SUM(importe_producido) AS importe_producido
         FROM ({sql_union}) AS c
         WHERE tipo_tiempo IS NOT NULL AND tipo_tiempo <> ''
         GROUP BY tipo_tiempo
      """

      sql_por_sitio = f"""
         SELECT sitio, SUM(importe_producido) AS importe_producido
         FROM ({sql_union}) AS c
         WHERE sitio <> 'SIN SITIO'
         GROUP BY sitio
         ORDER BY SUM(importe_producido) DESC
         LIMIT 15
      """

      sql_por_lider = f"""
         SELECT lider_proyecto, SUM(importe_producido) AS importe_producido
         FROM ({sql_union}) AS c
         GROUP BY lider_proyecto
         ORDER BY SUM(importe_producido) DESC
         LIMIT 10
      """

      sql_por_fecha_sitio = f"""
         SELECT fecha_produccion, sitio, SUM(importe_producido) AS importe_producido
         FROM ({sql_union}) AS c
         WHERE sitio <> 'SIN SITIO'
         GROUP BY fecha_produccion, sitio, _fecha_ord
         ORDER BY _fecha_ord ASC NULLS LAST, sitio ASC
      """

      fila_totales   = (ejecutar_query_sql(sql_resumen, params) or [{}])[0]
      fila_mejor     = (ejecutar_query_sql(sql_mejor_dia, params) or [{}])[0]
      fechas         = ejecutar_query_sql(sql_por_fecha, params) or []
      tipos          = ejecutar_query_sql(sql_por_tipo, params) or []
      sitios         = ejecutar_query_sql(sql_por_sitio, params) or []
      lideres        = ejecutar_query_sql(sql_por_lider, params) or []
      fecha_sitio    = ejecutar_query_sql(sql_por_fecha_sitio, params) or []

      return {
         "resumen": {
            "total_importe_producido": float(fila_totales.get("total_importe_producido") or 0),
            "proyectos_ejecutados":    int(fila_totales.get("proyectos_ejecutados") or 0),
            "dias_unicos":             int(fila_totales.get("dias_unicos") or 0),
            "mejor_dia_fecha":         fila_mejor.get("fecha_produccion", "—"),
            "mejor_dia_importe":       float(fila_mejor.get("importe_dia") or 0),
         },
         "por_fecha": [
            {
               "fecha":              r["fecha_produccion"],
               "importe_producido":  float(r["importe_producido"] or 0),
               "importe_programado": float(r["importe_programado"] or 0),
            }
            for r in fechas
         ],
         "por_tipo_tiempo": [
            {"tipo": r["tipo_tiempo"], "importe": float(r["importe_producido"] or 0)}
            for r in tipos
         ],
         "por_sitio": [
            {"sitio": r["sitio"], "importe": float(r["importe_producido"] or 0)}
            for r in sitios
         ],
         "por_lider": [
            {"lider": r["lider_proyecto"], "importe": float(r["importe_producido"] or 0)}
            for r in lideres
         ],
         "por_fecha_sitio": [
            {
               "fecha":   r["fecha_produccion"],
               "sitio":   r["sitio"],
               "importe": float(r["importe_producido"] or 0),
            }
            for r in fecha_sitio
         ],
      }

   except Exception as error_agg:
      print(f"Error en fn_calcular_aggregados_info: {str(error_agg)}")
      return {}